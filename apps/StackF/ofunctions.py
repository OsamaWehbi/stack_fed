import itertools

import numpy as np
from sklearn.cluster import DBSCAN
import openpyxl
import ast
from apps.StackF.fed_stack import Client
from src.apis.extensions import Dict
from collections import Counter
import copy
import random
import math


# this file include
# 1-shanon
# 2-dbscan
# 3-readexcel
# 4- compute combination
# 5- min_max for total data / deleted
# 6- min_max for the data dictionary
# 7- gen_pay payment from followers to leaders
# 8- compute_utility_follower
# 9- compute_utility_leader

def shanon(dn_val_list):
    # list of values
    values = dn_val_list
    # calculate the proportion of each value
    proportions = np.array(values) / np.sum(values)

    # calculate the Shannon diversity index
    shannon_index = -np.sum(proportions * np.log(proportions))
    # return results
    return shannon_index


def dbscan(dc_client):
    # list of clients of dictionary type
    clients = dc_client
    # convert the client data containers to a numpy array
    X = np.array([[client.cpu, client.ram, client.bandwidth, client.c_dd] for client in clients])

    # create a DBSCAN object and fit the data
    dbscan = DBSCAN(eps=30, min_samples=2)
    # dbscan = DBSCAN(eps=20, min_samples=2)
    dbscan.fit(X)

    # get the labels for the data points
    labels = dbscan.labels_

    # print the labels
    # print(labels)

    # separate the clients into two groups based on the labels
    group1 = [clients[i] for i in range(len(clients)) if labels[i] == 0]
    group2 = [clients[i] for i in range(len(clients)) if labels[i] == 1 or labels[i] == 2]
    # group1 = [clients[i] for i in range(len(clients)) if labels[i] == -1]
    # group2 = [clients[i] for i in range(len(clients)) if labels[i] == 0 or labels[i] == 2]

    # print the groups
    # print("Group 1:")
    # for client in group1:
    #     print(client.cpu, client.ram, client.bandwidth, client.c_dd)
    #
    # print("Group 2:")
    # for client in group2:
    #     print(client.cpu, client.ram, client.bandwidth, client.c_dd)

    return group1, group2


def readexcel(filename='C:/Users/Osama.Wehbi/Desktop/StackFed/apps/StackF/DATA.xlsx'):
    # Load the Excel file
    workbook = openpyxl.load_workbook(filename)

    # Select the sheet named "Clients"
    sheet = workbook['Clients']

    # Get the maximum row number and column letter
    max_row = sheet.max_row
    max_col = openpyxl.utils.get_column_letter(sheet.max_column)

    # Read the attribute names from the first row
    attributes = {}
    # for col in range(1, sheet.max_column + 1):
    #     attribute_name = sheet.cell(row=1, column=col).value
    #     attributes[attribute_name] = col

    # Create client objects based on the attribute values in each row
    clients = []
    #  list of dictionaries to be normalized
    dicts = []
    for row in range(1, sheet.max_row + 1):
        c_id = sheet.cell(row=row, column=1).value
        cpu = int(sheet.cell(row=row, column=2).value)
        ram = int(sheet.cell(row=row, column=3).value)
        bandwidth = int(sheet.cell(row=row, column=4).value)
        dn_data = ast.literal_eval(sheet.cell(row=row, column=5).value)
        c_dd = float(shanon(list(dn_data.values())))
        # print(c_id, cpu, ram, band, dn_data, c_dd)
        client = Client(c_id, cpu, ram, bandwidth, dn_data, c_dd)
        dicts.append(dn_data)
        clients.append(client)
    dicts = normalize_dicts(copy.deepcopy(dicts))
    for i in range(len(clients)):
        clients[i].c_dn_data_nor = dicts[i]
        # clients[i].c_dd = float(shanon(list(dicts[i].values())))

    # Do something with the list of client objects
    # print(clients)
    return clients


def compute_com(leaders, followers):
    total = 0
    for i in leaders:
        total += sum(list(i.c_dn_data_nor.values()))
    for i in followers:
        total += sum(list(i.c_dn_data_nor.values()))
    for leader in leaders:
        for follower in followers:
            Ac = float(len(list(set(list(leader.c_dn_data_nor.keys()) + list(follower.c_dn_data_nor.keys())))) / 10)
            Aq = (sum(list(leader.c_dn_data_nor.values())) + sum(list(follower.c_dn_data_nor.values())))
            val = dict(Counter(leader.dn_data) + Counter(follower.dn_data))
            Ad = shanon(list(val.values()))
            leader.com_dic[follower.c_id] = 0.33 * Ac * 0.33 * Aq * 0.33 * Ad


# def min_max(clients):
#     data = []
#     for i in clients:
#         data.append(sum(i.dn_data.values()))
#     min_val = min(data)
#     max_val = max(data)
#     for i in clients:
#         print(sum(list(i.dn_data.values())))
#         i.c_n_data_val = 100 * (sum(i.dn_data.values()) - min_val) / (max_val - min_val)
#         print(i.c_n_data_val)
#         print("hello")


def normalize_dicts(dicts):
    # print(dicts)
    val_list = []
    for i in dicts:
        val_list += list(i.values())
    # print(val_list)
    keys = set().union(*dicts)  # get the unique keys in all dictionaries
    for key in keys:
        values = [d.get(key, 0) for d in dicts]  # get the values for the current key in all dictionaries
        min_val = min(val_list)
        max_val = max(val_list)
        # print(min_val)
        # print(max_val)
        for d in dicts:
            if key in d:
                d[key] = round(100 * (d[key] - min_val) / (max_val - min_val))
    # print(dicts)
    return dicts


def gen_pay(followers, leaders):
    for follower in followers:
        # follower.f_l_pay = {leader.c_id: random.randint(10, 40) for leader in leaders}
        # print(leaders[0].com_dic["IOT2"], leaders[0].c_ds)
        follower.f_l_pay = {leader.c_id: (leader.com_dic[follower.c_id] - leader.c_ds) for leader in leaders}


def gen_pay_leader(followers, leaders):
    for leader in leaders:
        # follower.f_l_pay = {leader.c_id: random.randint(10, 40) for leader in leaders}
        # print(leaders[0].com_dic["IOT2"], leaders[0].c_ds)
        leader.l_f_pay_dic = {follower.c_id: (leader.com_dic[follower.c_id] - follower.c_ds) for follower in followers}


def compute_utility_follower(followers, leaders):
    for leader in leaders:
        total = 0
        # get the total payments to the specific leaders from all the followers
        for follower in followers:
            total += follower.f_l_pay[leader.c_id]
        # compute the utility between each follower and the leaders
        for follower in followers:
            follower.u_f[leader.c_id] = abs(((follower.f_l_pay[leader.c_id] / total) * (
                    leader.com_dic[follower.c_id] - follower.c_ds)) - follower.f_l_pay[leader.c_id])


def compute_utility_leader(followers, leaders):
    for leader in leaders:
        leader.u_l = 0
        for follower in followers:
            # print(leader.com_dic[follower.c_id], leader.c_ds, follower.f_l_pay[leader.c_id])
            leader.u_l += (leader.com_dic[follower.c_id] - leader.c_ds) + follower.f_l_pay[leader.c_id]


def compute_tot(followers, leader):
    total = 0
    # get the total payments to the specific leaders from all the followers
    for follower in followers:
        total += follower.f_l_pay[leader.c_id]
    return total


def compute_p_opt(followers, leaders):
    # for leader in leaders:
    #     for follower in followers:
    #         follower.eq18 = math.sqrt()
    for follower in followers:
        # for leader in leaders:
        #     print(leader.com_dic[follower.c_id], compute_tot(followers, leader))
        # print(leaders[0].l_f_pay_dic[follower.c_id], sum(leaders[0].l_f_pay_dic.values()), leaders[0].l_f_pay_dic[follower.c_id])
        follower.eq18_dic = {leader.c_id: abs(
            (math.sqrt(leader.l_f_pay_dic[follower.c_id] * (
                    sum(leader.l_f_pay_dic.values()) - leader.l_f_pay_dic[follower.c_id]))) - (
                    sum(leader.l_f_pay_dic.values()) - leader.l_f_pay_dic[follower.c_id]))
            for leader in leaders}


def get_combinations(list1, list2):
    combinations = []
    for item in list1:
        # leader1
        for combination in itertools.combinations(list2, item.c_q):
            combinations.append((item, *combination))
    return combinations


def com_final(combinations):
    com_values = []
    # counter = 0
    for com in combinations:
        value = 0
        for follower in com[1:]:
            value += follower.f_l_pay[com[0].c_id] + follower.eq18_dic[com[0].c_id]
        com_values.append(value)
        # counter += 1
    return com_values


def sorting_for_selec(combinations, com_values):
    # print(com_values)
    # print(combinations)
    numbers = com_values
    tuples = combinations
    # Combine the two lists into pairs
    combined = list(zip(numbers, tuples))

    # Sort the pairs based on the numbers
    sorted_combined = sorted(combined, key=lambda x: x[0], reverse=True)

    # Separate the pairs back into two separate lists
    sorted_numbers, sorted_tuples = zip(*sorted_combined)
    return list(sorted_tuples), list(sorted_numbers)


def delete_tuples_with_objects(tuple_list, object_list):
    updated_list = [t for t in tuple_list if not any(obj in t for obj in object_list)]
    return updated_list


# def sel_fun(combinations, leaders):
#     # hold the selected clients
#     selected = []
#     for com in combinations:
#         if com[0].c_s != 1:
#             com[0].c_s = 1
#             selected.append(com[0])
#             for follower in com[1:]:
#                 follower.c_s = 1
#                 selected.append(follower)
#             # combinations = [com for com in combinations if not any(client in selected for client in com)]
#         combinations = delete_tuples_with_objects(combinations, selected)
#             # print("this is the new commm", combinations)
#     return selected

def sel_fun(combinations, leaders):
    # hold the selected clients
    selected = []
    temp = []
    for _ in range(len(leaders)):
        temp, combinations = inter_select(combinations)
        selected += temp

    return selected


def inter_select(combinations):
    combinations = combinations
    selected = []
    if combinations[0][0].c_s != 1:
        combinations[0][0].c_s = 1
        selected.append(combinations[0][0])
        for follower in combinations[0][1:]:
            follower.c_s = 1
            selected.append(follower)
    combinations = delete_tuples_with_objects(combinations, selected)
    return selected, combinations


def sel_fun_opt(combinations):
    # Create a set to store the selected clients
    selected = set()

    # Iterate over combinations
    for com in combinations:
        if com[0].c_s != 1:
            # Select the leader
            com[0].c_s = 1
            selected.add(com[0])

            # Select the followers
            for follower in com[1:]:
                follower.c_s = 1
                selected.add(follower)

        # Filter out the selected clients from combinations
        combinations = [com for com in combinations if not any(client in selected for client in com)]

    return list(selected)


def transfer_to_id_list(tuple_list):
    id_list = [(obj.c_id for obj in tup) for tup in tuple_list]
    return id_list
