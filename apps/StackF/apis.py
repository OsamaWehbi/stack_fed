import sqlite3

import sqlite3
import pandas as pd
from openpyxl import Workbook
import json

def print_db(file_path="../StackF/perf.db"):
    # Connect to the database file
    conn = sqlite3.connect(file_path)

    # Create a cursor object to execute SQL queries
    cursor = conn.cursor()

    # Execute an SQL query to fetch data
    cursor.execute("SELECT * FROM mnist")

    # Fetch all rows from the query result
    rows = cursor.fetchall()

    # Process the fetched rows
    for row in rows:
        print(row[0], row[1])

    # Close the cursor and connection
    cursor.close()
    conn.close()


def dumb_db(tab_id, file_path="../StackF/perf.db"):
    # Connect to the database file
    conn = sqlite3.connect(file_path)

    # Execute a query to fetch data from specific columns
    cursor = conn.cursor()
    cursor.execute(f"SELECT * FROM {tab_id}")
    rows = cursor.fetchall()

    # Create a pandas DataFrame from the fetched rows
    df = pd.DataFrame(rows)

    # Save the DataFrame to an Excel file
    df.to_excel(f'{tab_id}.xlsx', index=False)

    # Close the cursor and connection
    cursor.close()
    conn.close()


def write_ds(clients):
    # Create a new workbook and select the active sheet
    workbook = Workbook()
    sheet = workbook.active
    sheet['A1'] = 'Device'
    sheet['B1'] = 'Ds'
    for i, client in enumerate(clients, 2):
        # Write data to the sheet
        sheet['A' + str(i)] = str(client.c_id)
        sheet['B' + str(i)] = str(client.c_ds)
    # Save the workbook to a file
    workbook.save('ds.xlsx')


def final_val(selected):
    # Create a new workbook and select the active sheet
    workbook = Workbook()
    sheet = workbook.active

    # Define the list of objects

    # Define the headers and write them to the sheet
    headers = ['c_id', 'cpu', 'ram', 'c_ds', 'com_dic', 'f_l_pay', 'u_f', 'u_l', 'l_f_pay_dic', 'c_q', 'c_s', 'eq18_dic']  # Replace with your attribute names

    for col_num, header in enumerate(headers, 1):
        sheet.cell(row=1, column=col_num).value = header

    # Write the object attributes to the sheet
    for row_num, obj in enumerate(selected, 2):
        sheet.cell(row=row_num, column=1).value = obj.c_id  # Replace with your attribute names
        sheet.cell(row=row_num, column=2).value = obj.cpu
        sheet.cell(row=row_num, column=3).value = obj.ram
        sheet.cell(row=row_num, column=4).value = obj.c_ds
        sheet.cell(row=row_num, column=5).value = json.dumps(obj.com_dic)
        sheet.cell(row=row_num, column=6).value = json.dumps(obj.f_l_pay)
        sheet.cell(row=row_num, column=7).value = json.dumps(obj.u_f)
        sheet.cell(row=row_num, column=8).value = obj.u_l
        sheet.cell(row=row_num, column=9).value = json.dumps(obj.l_f_pay_dic)
        sheet.cell(row=row_num, column=10).value = obj.c_q
        sheet.cell(row=row_num, column=11).value = obj.c_s
        sheet.cell(row=row_num, column=12).value = json.dumps(obj.eq18_dic)

    # Save the workbook to a file
    workbook.save('all_values.xlsx')

