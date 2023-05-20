import numpy as np
import openpyxl
from scipy.stats import entropy

# create an example list of values
values = [99,82,79]

# calculate the proportion of each value
proportions = np.array(values) / np.sum(values)

# calculate the Shannon diversity index
shannon_index = -np.sum(proportions * np.log(proportions))

# print the results
print("Values:", values)
print("Proportions:", proportions)
print("Shannon Diversity Index:", shannon_index)

# Load the Excel file
workbook = openpyxl.load_workbook('C:/Users/Osama.Wehbi/Desktop/StackFed/apps/StackF/DATA.xlsx')

# Select the sheet named "Clients"
sheet = workbook['Clients']

# Get the maximum row number and column letter
max_row = sheet.max_row
max_col = openpyxl.utils.get_column_letter(sheet.max_column)

# Read the attribute names from the first row
attributes = {}
# for col in range(1, sheet.max_column + 1):
#     attribute_name = sheet.cell(row=1, column=col).value
#     attributes[attribute_name] = col

# Create client objects based on the attribute values in each row
clients = []
for row in range(1, sheet.max_row):
    c_id = sheet.cell(row=row, column=1).value
    cpu = sheet.cell(row=row, column=2).value
    ram = sheet.cell(row=row, column=3).value
    band = sheet.cell(row=row, column=4).value
    dn_data = sheet.cell(row=row, column=5).value
    c_dd = sheet.cell(row=row, column=6).value
    print(c_id,cpu,ram,band,dn_data,c_dd)
    # client = Client(cpu, ram, band)
    # clients.append(client)

# Do something with the list of client objects
# print(clients)
